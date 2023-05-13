import openpyxl
import requests
import json
from io import BytesIO

XLSX_URL: str = r'https://docs.google.com/spreadsheets/d/1W5znKQixeRJeg_IwiLoUFN8cqcStC44L/export?format=xlsx'
OUTPUT_FILE: str = "frontend/dist/schedule.json"
TABLE_START: str = 'A5'
DATE_CELL: str = "A3"


def download_xlsx(url: str) -> BytesIO:
    # Downloading the file and return it as a file like object
    return BytesIO(requests.get(XLSX_URL).content)


def get_col(worksheet, starting_cell):
    return [cell.value for cell in list(worksheet[starting_cell.column_letter]) if cell.row > starting_cell.row]


def get_row(worksheet, starting_cell):
    return {cell.value: cell for cell in list(worksheet[starting_cell.row]) if cell.column > starting_cell.column}


def main():
    xlsx_file = download_xlsx(XLSX_URL)
    ws = openpyxl.load_workbook(filename=xlsx_file).worksheets[0]
    classes = get_row(worksheet=ws, starting_cell=ws[TABLE_START])

    hours = get_col(worksheet=ws, starting_cell=ws[TABLE_START])

    schedule = {"date": ws[DATE_CELL].value}
    for cur_class in classes:
        class_schedule = get_col(
            worksheet=ws, starting_cell=classes[cur_class])
        schedule[cur_class] = [[hour, lesson]
                               for hour, lesson in zip(hours, class_schedule)]

    with open(OUTPUT_FILE, "w") as outputFile:
        json.dump(schedule, outputFile)


if __name__ == "__main__":
    main()
