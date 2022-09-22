import requests
import openpyxl
import logging
import functions_framework
import json


# add conf.json
XLSX_URL: str = r'https://docs.google.com/spreadsheets/d/1Eq3CQgjbVGXv0DXgsztmfax6JWu8iYH3/export?format=xlsx'
XLSX_FILENAME: str = 'schedule.xlsx'
DEFAULT_TABLE_START: str = 'A5'


def download_xlsx(url: str) -> None:
    req = requests.get(url, allow_redirects=True)
    with open(XLSX_FILENAME, 'wb') as f:
        f.write(req.content)
    print(f'{XLSX_FILENAME} saved.')

    
def get_col(worksheet, starting_cell):
    return [cell.value for cell in list(worksheet[starting_cell.column_letter]) if cell.row > starting_cell.row]
  
  
def get_row(worksheet, starting_cell):
    return {cell.value: cell for cell in list(worksheet[starting_cell.row]) if cell.column > starting_cell.column}

  
@functions_framework.http
def get_req(request):
    #download_xlsx(XLSX_URL)
    ws = openpyxl.load_workbook(filename=XLSX_FILENAME).worksheets[0]
        
    table_start: str = request.args.get('table_start') or DEFAULT_TABLE_START
    classes = get_row(worksheet=ws, starting_cell=ws[table_start])
    if request.args.get('getClasses') == "true":
        # /?getClasses=true
        response = json.dumps(list(classes.keys()))
    else:
        chosen_class: str = request.args.get('class')
        print(chosen_class)
        hours = get_col(worksheet=ws, starting_cell=ws[DEFAULT_TABLE_START]) 
        class_schedule = get_col(worksheet=ws, starting_cell=classes[chosen_class])
        response = {hour: lesson for hour, lesson in zip(hours, class_schedule)}
    return response, 200, {'Access-Control-Allow-Origin': '*'}
