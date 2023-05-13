import openpyxl
import requests
from flask import Flask, make_response

# add conf.json
XLSX_URL: str = r'https://docs.google.com/spreadsheets/d/1W5znKQixeRJeg_IwiLoUFN8cqcStC44L/export?format=xlsx'
XLSX_FILENAME: str = '/tmp/schedule.xlsx'
TABLE_START: str = 'A5'
DATE_CELL: str = "A3"

app = Flask(__name__)

@app.route('/api/schedule')
def schedule():
    resp = make_response(create_schedule())
    resp.headers["Cache-Control"] = "max-age=0, s-maxage=1000"
    return resp

    
def download_xlsx(url: str) -> None:
    with open(XLSX_FILENAME, "wb") as f:
        f.write(requests.get(XLSX_URL).content)


def get_col(worksheet, starting_cell):
    return [cell.value for cell in list(worksheet[starting_cell.column_letter]) if cell.row > starting_cell.row]


def get_row(worksheet, starting_cell):
    return {cell.value: cell for cell in list(worksheet[starting_cell.row]) if cell.column > starting_cell.column}


def create_schedule() -> dict[str, str]:
    download_xlsx(XLSX_URL)
    ws = openpyxl.load_workbook(filename=XLSX_FILENAME).worksheets[0]
    classes = get_row(worksheet=ws, starting_cell=ws[TABLE_START])

    hours = get_col(worksheet=ws, starting_cell=ws[TABLE_START])

    schedule = {"date": ws[DATE_CELL].value}
    for cur_class in classes:
        class_schedule = get_col(
            worksheet=ws, starting_cell=classes[cur_class])
        schedule[cur_class] = [[hour, lesson]
                               for hour, lesson in zip(hours, class_schedule)]

    return schedule

